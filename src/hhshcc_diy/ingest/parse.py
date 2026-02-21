"""Parse CMS DIY Excel tables into a ModelData object."""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd

from hhshcc_diy.model.data import ModelData

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_hcc_numbers(text: str) -> list[str]:
    """Extract HCC numbers from SAS-like definition text.

    Handles patterns like:
      HHS_HCC019, HHS_HCC035_1, HHS_HCC161_2
    Returns: ["019", "035_1", "161_2"]
    """
    return re.findall(r"HHS_HCC(\d{3}(?:_\d+)?)", text)


def _extract_group_name(text: str) -> str | None:
    """Extract group variable name from definition text like 'G01 = 1'."""
    m = re.search(r"\b(G\d+\w*)\s*=\s*1", text)
    return m.group(1) if m else None


def _extract_rxc_number(text: str) -> str | None:
    """Extract RXC number like 'RXC_01' from text."""
    m = re.search(r"RXC_(\d+)", text)
    return m.group(1) if m else None


def _normalize_hcc_key(raw: str) -> str:
    """Normalize an HCC key to zero-padded format.

    '1' → '001', '35_1' → '035_1', '161_2' → '161_2'
    """
    if "_" in raw:
        base, suffix = raw.split("_", 1)
        return f"{base.zfill(3)}_{suffix}"
    return raw.zfill(3)


def _format_cc(x) -> str | None:
    """Format a CC value from Excel, preserving split suffixes.

    Excel stores CC numbers as floats. Plain CCs like 35.0 → "035".
    Split CCs like 35.1 → "035_1", 35.2 → "035_2".
    """
    if pd.isna(x) or str(x).strip() == "":
        return None
    val = float(x)
    int_part = int(val)
    frac = round(val - int_part, 1)
    if frac == 0.0:
        return str(int_part).zfill(3)
    else:
        suffix = int(round(frac * 10))
        return f"{str(int_part).zfill(3)}_{suffix}"


def _normalize_sheet_name(wb_sheets: list[str], target: str) -> str:
    """Find a sheet name matching target, tolerating trailing spaces."""
    for s in wb_sheets:
        if s.strip() == target.strip():
            return s
    raise KeyError(f"Sheet '{target}' not found in workbook. Available: {wb_sheets}")


def _read_sheet(excel_path: Path, sheet_name: str, header_row: int,
                skip_footer: int = 0) -> pd.DataFrame:
    """Read a sheet with the given header row (0-indexed)."""
    df = pd.read_excel(
        excel_path, sheet_name=sheet_name,
        header=header_row, engine="openpyxl",
    )
    if skip_footer > 0:
        df = df.iloc[:-skip_footer]
    return df


# ---------------------------------------------------------------------------
# Table-specific parsers
# ---------------------------------------------------------------------------

def _parse_table3(excel_path: Path, sheet_name: str, benefit_year: int) -> pd.DataFrame:
    """Parse Table 3: ICD-10 → CC Crosswalk."""
    df = _read_sheet(excel_path, sheet_name, header_row=3)

    # Standardize column names — they vary slightly by year
    col_map = {}
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if col_lower == "icd10":
            col_map[col] = "ICD10"
        elif "cc" == col_lower and "second" not in col_lower and "third" not in col_lower:
            col_map[col] = "CC"
        elif "second cc" in col_lower or col_lower == "second cc":
            col_map[col] = "SECOND_CC"
        elif "third cc" in col_lower or col_lower == "third cc":
            col_map[col] = "THIRD_CC"
        elif "cc age split" in col_lower:
            col_map[col] = "CC_AGE_SPLIT"
        elif "cc sex split" in col_lower:
            col_map[col] = "CC_SEX_SPLIT"

    # Find the CY-level MCE columns (combined FY columns)
    for col in df.columns:
        col_str = str(col).strip()
        if re.search(r"CY\d{4}.*MCE Age", col_str, re.IGNORECASE):
            col_map[col] = "MCE_AGE_CONDITION"
        elif re.search(r"CY\d{4}.*MCE Sex", col_str, re.IGNORECASE):
            col_map[col] = "MCE_SEX_CONDITION"

    # If no CY-level MCE columns, fall back to FY-specific ones
    if "MCE_AGE_CONDITION" not in col_map.values():
        for col in df.columns:
            col_str = str(col).strip()
            # Use the first FY MCE Age column
            if re.search(r"FY\d{4}.*MCE Age", col_str, re.IGNORECASE):
                col_map[col] = "MCE_AGE_CONDITION"
                break
    if "MCE_SEX_CONDITION" not in col_map.values():
        for col in df.columns:
            col_str = str(col).strip()
            if re.search(r"FY\d{4}.*MCE Sex", col_str, re.IGNORECASE):
                col_map[col] = "MCE_SEX_CONDITION"
                break

    # Find Code Valid columns
    for col in df.columns:
        col_str = str(col).strip()
        if re.search(r"Code Valid.*FY\d{4}", col_str, re.IGNORECASE):
            fy_match = re.search(r"FY(\d{4})", col_str)
            if fy_match:
                fy = int(fy_match.group(1))
                if f"CODE_VALID_FY{fy}" not in col_map.values():
                    col_map[col] = f"CODE_VALID_FY{fy}"

    df = df.rename(columns=col_map)

    # Keep only mapped columns plus CC columns
    keep = [c for c in df.columns if c in col_map.values() or c in ("CC", "SECOND_CC", "THIRD_CC", "Obs", "ICD10 Label", "Footnote")]

    # Ensure we have the CC column — it might be the 14th column (N) by position
    if "CC" not in df.columns:
        # Find column by position — CC is column N (index 13) per the PDF
        orig_cols = list(df.columns)
        for i, col in enumerate(orig_cols):
            col_str = str(col).strip().lower()
            if col_str == "cc":
                df = df.rename(columns={col: "CC"})
                break

    # Drop rows where ICD10 is NaN (notes section at bottom)
    df = df.dropna(subset=["ICD10"])

    # Remove notes rows (ICD10 starts with "Notes" or is non-alphanumeric)
    df = df[df["ICD10"].astype(str).str.match(r"^[A-Z]\d", na=False)]

    # Convert CC columns to string, preserving split suffixes (e.g., 35.1 → "035_1")
    for cc_col in ["CC", "SECOND_CC", "THIRD_CC"]:
        if cc_col in df.columns:
            df[cc_col] = df[cc_col].apply(_format_cc)

    logger.info("Table 3: %d ICD-10 codes loaded", len(df))
    return df


def _parse_table4(excel_path: Path, sheet_name: str) -> tuple[dict[str, list[str]], list[str]]:
    """Parse Table 4: HCC Hierarchies. Returns (hierarchy_dict, all_hccs)."""
    df = _read_sheet(excel_path, sheet_name, header_row=2)

    # Columns: Obs, V08 HCC (or V07 HCC), Set to 0 HCCs, HCC Label
    # Find the HCC column
    hcc_col = None
    zero_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if "hcc" in col_str and "set to" not in col_str and "label" not in col_str:
            hcc_col = col
        elif "set to 0" in col_str:
            zero_col = col

    if hcc_col is None or zero_col is None:
        raise ValueError(f"Cannot identify HCC/hierarchy columns in Table 4: {list(df.columns)}")

    df = df.dropna(subset=[hcc_col])

    hierarchy = {}
    all_hccs = []

    for _, row in df.iterrows():
        hcc = _normalize_hcc_key(str(row[hcc_col]).strip())
        all_hccs.append(hcc)

        zero_val = row[zero_col]
        if pd.isna(zero_val) or str(zero_val).strip() == "":
            hierarchy[hcc] = []
        else:
            # Parse comma-separated HCC numbers, which may include ranges or splits
            zero_str = str(zero_val).strip()
            zeros = [_normalize_hcc_key(z.strip()) for z in re.split(r"[,;\s]+", zero_str) if z.strip()]
            hierarchy[hcc] = zeros

    logger.info("Table 4: %d HCCs with hierarchy rules", len(hierarchy))
    return hierarchy, all_hccs


def _parse_table5_age_sex(excel_path: Path, sheet_name: str) -> dict[str, list[str]]:
    """Parse Table 5: Age-Sex Variables. Returns {model: [variable_names]}."""
    df = _read_sheet(excel_path, sheet_name, header_row=2)

    # Columns: Model, Variable, Description, Variable Used in Risk Score Formula?, Definition
    var_col = None
    model_col = None
    used_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if col_str == "variable":
            var_col = col
        elif col_str == "model":
            model_col = col
        elif "used in risk score" in col_str:
            used_col = col

    result: dict[str, list[str]] = {"Adult": [], "Child": [], "Infant": []}
    current_model = None

    for _, row in df.iterrows():
        if pd.notna(row.get(model_col)):
            current_model = str(row[model_col]).strip()

        if current_model is None:
            continue

        var_name = row.get(var_col)
        used = row.get(used_col)
        if pd.notna(var_name) and pd.notna(used) and str(used).strip().lower() == "yes":
            var_name = str(var_name).strip()
            # Only include age-sex demographic variables (MAGE_LAST_*, FAGE_LAST_*, AGE0_MALE, AGE1_MALE)
            if any(var_name.startswith(p) for p in ("MAGE_LAST_", "FAGE_LAST_", "AGE0_MALE", "AGE1_MALE")):
                for model in result:
                    if model.lower() in current_model.lower() or current_model.lower().startswith(model.lower()):
                        result[model].append(var_name)

    # Handle combined model entries like "Adult, Child, Infant"
    # Re-parse more carefully
    result = {"Adult": [], "Child": [], "Infant": []}
    current_models = []
    for _, row in df.iterrows():
        if pd.notna(row.get(model_col)):
            m = str(row[model_col]).strip()
            current_models = [s.strip() for s in m.split(",")]

        var_name = row.get(var_col)
        used = row.get(used_col)
        if pd.notna(var_name) and pd.notna(used) and str(used).strip().lower() == "yes":
            var_name = str(var_name).strip()
            for model in current_models:
                if model in result:
                    result[model].append(var_name)

    logger.info("Table 5 age-sex vars: Adult=%d, Child=%d, Infant=%d",
                len(result["Adult"]), len(result["Child"]), len(result["Infant"]))
    return result


def _parse_variable_table(excel_path: Path, sheet_name: str, model_name: str) -> dict:
    """Parse Tables 6/7/8: variable definitions for a model segment.

    Returns a dict with parsed components:
    - hcc_groups: {group_name: [hcc_numbers]}
    - rxc_interactions: [dict] (adult only)
    - severe_triggers: [var_names]
    - transplant_triggers: [var_names]
    - severe_hcc_count_bins: [(var_name, min, max_or_none)]
    - transplant_hcc_count_bins: [(var_name, min, max_or_none)]
    - enrollment_duration_vars: [(var_name, month)] (adult only)
    - hcc_cnt_exclude: str|None
    - infant_severity_levels: {var: [hccs]} (infant only)
    - infant_maturity_categories: [dict] (infant only)
    - infant_maturity_severity_interactions: [(var, maturity, severity)] (infant only)
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()

    result = {
        "hcc_groups": {},
        "rxc_interactions": [],
        "severe_triggers": [],
        "transplant_triggers": [],
        "severe_hcc_count_bins": [],
        "transplant_hcc_count_bins": [],
        "enrollment_duration_vars": [],
        "hcc_cnt_exclude": None,
        "infant_severity_levels": {},
        "infant_maturity_categories": [],
        "infant_maturity_severity_interactions": [],
    }

    # Parse row by row, tracking context
    current_var = None
    current_section = None  # group, rxc_interaction, severe, transplant, etc.

    for i, row in enumerate(rows):
        if i < 2:  # Skip title and blank rows
            continue

        # Column mapping: A=Model, B=Variable, C=Description, D=Used, E=Definition
        model_val = row[0] if len(row) > 0 else None
        var_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        desc = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
        used = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
        defn = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None

        if defn is None or defn == "None" or defn == "":
            continue

        # Detect variable type from variable name and context
        if var_name and var_name != "None":
            current_var = var_name

            # RXC standalone variables (RXC_01 through RXC_10) — skip, these are just definitions
            if re.match(r"^RXC_\d+$", var_name):
                current_section = "rxc_standalone"
                continue

            # RXC interaction variables
            if re.match(r"^RXC_\d+_X_", var_name):
                current_section = "rxc_interaction"
                # Parse the interaction
                rxc_match = re.match(r"^RXC_(\d+)_X_", var_name)
                rxc_num = rxc_match.group(1) if rxc_match else None

                # Extract HCC numbers from the definition
                hcc_nums = _extract_hcc_numbers(defn)

                # Detect complex logic (AND between sets)
                if "_AND_" in var_name:
                    # Complex: two sets of HCCs, both must have at least one
                    # e.g. RXC_09_X_HCC056_057_AND_048_041
                    parts = var_name.split("_AND_")
                    set1_hccs = _extract_hcc_numbers(parts[0])
                    set2_hccs = _extract_hcc_numbers(parts[1].replace("RXC_" + rxc_num + "_X_", ""))
                    # Actually, parse from the definition which is more reliable
                    # "if RXC_09 = 1 and (HHS_HCC056 = 1 or HHS_HCC057 = 1) and (HHS_HCC048 = 1 or HHS_HCC041 = 1)"
                    and_parts = re.split(r"\)\s*and\s*\(", defn)
                    if len(and_parts) >= 2:
                        set1 = _extract_hcc_numbers(and_parts[0])
                        set2 = _extract_hcc_numbers(and_parts[1])
                        # Remove the RXC's own HCCs if accidentally captured
                        result["rxc_interactions"].append({
                            "name": var_name,
                            "rxc": rxc_num,
                            "logic": "all_sets",
                            "hcc_sets": [set1, set2],
                        })
                    else:
                        result["rxc_interactions"].append({
                            "name": var_name,
                            "rxc": rxc_num,
                            "logic": "any",
                            "hccs": hcc_nums,
                        })
                else:
                    result["rxc_interactions"].append({
                        "name": var_name,
                        "rxc": rxc_num,
                        "logic": "any",
                        "hccs": hcc_nums,
                    })
                continue

            # HCC groups (G01, G02B, etc.)
            if re.match(r"^G\d+", var_name):
                current_section = "group"
                group_name = var_name
                hccs = _extract_hcc_numbers(defn)
                result["hcc_groups"][group_name] = hccs
                continue

            # HCC_CNT
            if var_name == "HCC_CNT":
                current_section = "hcc_cnt"
                # Check for exclusion: "- HHS_HCC022" in definition
                excl = re.search(r"-\s*HHS_HCC(\d{3})", defn)
                result["hcc_cnt_exclude"] = excl.group(1) if excl else None
                continue

            # SEVERE indicator
            if var_name == "SEVERE":
                current_section = "severe"
                triggers = _extract_hcc_numbers(defn)
                # Also check for group triggers (G13, G14, etc.)
                group_triggers = re.findall(r"\b(G\d+\w*)\s*=\s*1", defn)
                result["severe_triggers"] = [f"HHS_HCC{h}" for h in triggers] + group_triggers
                continue

            # TRANSPLANT indicator
            if var_name == "TRANSPLANT":
                current_section = "transplant"
                triggers = _extract_hcc_numbers(defn)
                group_triggers = re.findall(r"\b(G\d+\w*)\s*=\s*1", defn)
                result["transplant_triggers"] = [f"HHS_HCC{h}" for h in triggers] + group_triggers
                continue

            # SEVERE_HCC_COUNT bins
            if var_name.startswith("SEVERE_HCC_COUNT"):
                current_section = "severe_count"
                # Parse bin: SEVERE_HCC_COUNT1, SEVERE_HCC_COUNT10PLUS, SEVERE_HCC_COUNT6_7
                suffix = var_name.replace("SEVERE_HCC_COUNT", "")
                if "PLUS" in suffix.upper():
                    min_val = int(suffix.replace("PLUS", "").replace("plus", ""))
                    result["severe_hcc_count_bins"].append((var_name, min_val, None))
                elif "_" in suffix:
                    parts = suffix.split("_")
                    result["severe_hcc_count_bins"].append((var_name, int(parts[0]), int(parts[1])))
                else:
                    val = int(suffix)
                    result["severe_hcc_count_bins"].append((var_name, val, val))
                continue

            # TRANSPLANT_HCC_COUNT bins
            if var_name.startswith("TRANSPLANT_HCC_COUNT"):
                current_section = "transplant_count"
                suffix = var_name.replace("TRANSPLANT_HCC_COUNT", "")
                if "PLUS" in suffix.upper():
                    min_val = int(suffix.replace("PLUS", "").replace("plus", ""))
                    result["transplant_hcc_count_bins"].append((var_name, min_val, None))
                elif "_" in suffix:
                    parts = suffix.split("_")
                    result["transplant_hcc_count_bins"].append((var_name, int(parts[0]), int(parts[1])))
                else:
                    val = int(suffix)
                    result["transplant_hcc_count_bins"].append((var_name, val, val))
                continue

            # Enrollment duration vars (HCC_ED1 through HCC_ED6)
            if re.match(r"^HCC_ED\d+$", var_name):
                current_section = "enrollment"
                month = int(var_name.replace("HCC_ED", ""))
                result["enrollment_duration_vars"].append((var_name, month))
                continue

            # Infant severity levels
            if var_name.startswith("IHCC_SEVERITY"):
                current_section = "infant_severity"
                hccs = _extract_hcc_numbers(defn)
                result["infant_severity_levels"][var_name] = hccs
                continue

            # Infant maturity categories
            if var_name.startswith("IHCC_") and var_name not in ("IHCC_SEVERITY5", "IHCC_SEVERITY4",
                                                                   "IHCC_SEVERITY3", "IHCC_SEVERITY2",
                                                                   "IHCC_SEVERITY1"):
                if var_name in ("IHCC_AGE1", "IHCC_EXTREMELY_IMMATURE", "IHCC_IMMATURE",
                                "IHCC_PREMATURE_MULTIPLES", "IHCC_TERM"):
                    current_section = "infant_maturity"
                    newborn_hccs = _extract_hcc_numbers(defn)
                    age_match = re.search(r"AGE_LAST\s*=\s*(\d+)", defn)
                    age = int(age_match.group(1)) if age_match else None
                    result["infant_maturity_categories"].append({
                        "name": var_name,
                        "age": age,
                        "newborn_hccs": newborn_hccs,
                    })
                continue

            # Infant maturity × severity interactions
            if re.match(r"(EXTREMELY_IMMATURE|IMMATURE|PREMATURE_MULTIPLES|TERM|AGE1)_X_SEVERITY\d", var_name.strip()):
                current_section = "infant_mat_sev"
                # Parse: EXTREMELY_IMMATURE_X_SEVERITY5
                m = re.match(r"(.+)_X_(SEVERITY\d)", var_name.strip())
                if m:
                    maturity = "IHCC_" + m.group(1)
                    severity = "IHCC_" + m.group(2)
                    # Handle AGE1 → IHCC_AGE1
                    result["infant_maturity_severity_interactions"].append(
                        (var_name.strip(), maturity, severity)
                    )
                continue

        else:
            # Continuation row — same variable, additional definition lines
            if current_section == "group" and defn:
                # More HCCs for the same group
                hccs = _extract_hcc_numbers(defn)
                group_name_from_defn = _extract_group_name(defn)
                if group_name_from_defn and group_name_from_defn in result["hcc_groups"]:
                    result["hcc_groups"][group_name_from_defn].extend(hccs)
                elif current_var and current_var in result["hcc_groups"]:
                    result["hcc_groups"][current_var].extend(hccs)

            elif current_section == "severe" and defn:
                triggers = _extract_hcc_numbers(defn)
                group_triggers = re.findall(r"\b(G\d+\w*)\s*=\s*1", defn)
                result["severe_triggers"].extend([f"HHS_HCC{h}" for h in triggers])
                result["severe_triggers"].extend(group_triggers)

            elif current_section == "transplant" and defn:
                triggers = _extract_hcc_numbers(defn)
                group_triggers = re.findall(r"\b(G\d+\w*)\s*=\s*1", defn)
                result["transplant_triggers"].extend([f"HHS_HCC{h}" for h in triggers])
                result["transplant_triggers"].extend(group_triggers)

            elif current_section == "infant_severity" and defn:
                hccs = _extract_hcc_numbers(defn)
                if current_var and current_var in result["infant_severity_levels"]:
                    result["infant_severity_levels"][current_var].extend(hccs)

    # Deduplicate HCC lists within groups (regex can match same HCC twice
    # from assignment and zeroing portions of the SAS pseudocode)
    for group_name in result["hcc_groups"]:
        result["hcc_groups"][group_name] = list(
            dict.fromkeys(result["hcc_groups"][group_name])
        )

    return result


def _parse_table9(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    """Parse Table 9: Model Factors."""
    df = _read_sheet(excel_path, sheet_name, header_row=2)

    # Standardize column names
    col_map = {}
    for col in df.columns:
        col_str = str(col).strip().lower()
        if col_str == "model":
            col_map[col] = "Model"
        elif col_str == "variable":
            col_map[col] = "Variable"
        elif "platinum" in col_str:
            col_map[col] = "Platinum"
        elif "gold" in col_str:
            col_map[col] = "Gold"
        elif "silver" in col_str:
            col_map[col] = "Silver"
        elif "bronze" in col_str:
            col_map[col] = "Bronze"
        elif "catastrophic" in col_str:
            col_map[col] = "Catastrophic"
        elif "used in risk" in col_str:
            col_map[col] = "Used"

    df = df.rename(columns=col_map)

    # Forward-fill the Model column
    df["Model"] = df["Model"].ffill()

    # Drop rows where Variable is NaN
    df = df.dropna(subset=["Variable"])

    # Strip variable names
    df["Variable"] = df["Variable"].astype(str).str.strip()

    # Keep only scoring variables
    keep_cols = ["Model", "Variable", "Platinum", "Gold", "Silver", "Bronze", "Catastrophic"]
    df = df[[c for c in keep_cols if c in df.columns]]

    # Drop rows with NaN coefficients (section headers, etc.)
    df = df.dropna(subset=["Platinum"])

    logger.info("Table 9: %d factor rows loaded", len(df))
    return df


def _parse_table10a(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    """Parse Table 10a: NDC → RXC crosswalk."""
    df = _read_sheet(excel_path, sheet_name, header_row=3)

    # Columns: RXC, RXC_Label, NDC
    rxc_col = None
    ndc_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if col_str == "rxc":
            rxc_col = col
        elif col_str == "ndc":
            ndc_col = col

    if rxc_col is None or ndc_col is None:
        raise ValueError(f"Cannot find RXC/NDC columns in Table 10a: {list(df.columns)}")

    df = df.dropna(subset=[ndc_col])
    df = df[[rxc_col, ndc_col]].rename(columns={rxc_col: "RXC", ndc_col: "NDC"})
    df["RXC"] = df["RXC"].ffill().astype(int).astype(str).str.zfill(2)
    df["NDC"] = df["NDC"].astype(str).str.strip().str.zfill(11)

    logger.info("Table 10a: %d NDC→RXC mappings", len(df))
    return df


def _parse_table10b(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    """Parse Table 10b: HCPCS → RXC crosswalk."""
    df = _read_sheet(excel_path, sheet_name, header_row=3)

    rxc_col = None
    hcpcs_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if col_str == "rxc":
            rxc_col = col
        elif col_str == "hcpcs":
            hcpcs_col = col

    if rxc_col is None or hcpcs_col is None:
        raise ValueError(f"Cannot find RXC/HCPCS columns in Table 10b: {list(df.columns)}")

    df = df.dropna(subset=[hcpcs_col])
    df = df[[rxc_col, hcpcs_col]].rename(columns={rxc_col: "RXC", hcpcs_col: "HCPCS"})
    df["RXC"] = df["RXC"].ffill().astype(int).astype(str).str.zfill(2)
    df["HCPCS"] = df["HCPCS"].astype(str).str.strip()

    logger.info("Table 10b: %d HCPCS→RXC mappings", len(df))
    return df


def _parse_table11(excel_path: Path, sheet_name: str) -> tuple[dict[str, list[str]], list[str]]:
    """Parse Table 11: RXC Hierarchies. Returns (hierarchy_dict, all_rxcs)."""
    df = _read_sheet(excel_path, sheet_name, header_row=2)

    rxc_col = None
    zero_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if "rxc" == col_str:
            rxc_col = col
        elif "set to 0" in col_str:
            zero_col = col

    if rxc_col is None:
        raise ValueError(f"Cannot find RXC column in Table 11: {list(df.columns)}")

    df = df.dropna(subset=[rxc_col])

    hierarchy = {}
    all_rxcs = []

    for _, row in df.iterrows():
        rxc = str(row[rxc_col]).strip().zfill(2)
        all_rxcs.append(rxc)
        if zero_col and pd.notna(row[zero_col]) and str(row[zero_col]).strip():
            zeros = [z.strip().zfill(2) for z in re.split(r"[,;\s]+", str(row[zero_col]).strip()) if z.strip()]
            hierarchy[rxc] = zeros
        else:
            hierarchy[rxc] = []

    logger.info("Table 11: %d RXCs", len(all_rxcs))
    return hierarchy, all_rxcs


def _parse_table13(excel_path: Path, sheet_names: list[str]) -> dict[int, float]:
    """Parse Table 13: CSR Indicators. Returns csr_factors.

    Table 13 maps HIOS Variant IDs to CSR Indicators and factors.
    We extract the unique CSR Indicator → Factor mapping.
    The CSR factor is used solely as a multiplier in Step 5; it does not
    influence which metal level's coefficients are used for scoring.
    """
    if "Table 13" not in sheet_names:
        return _default_csr_factors()

    df = _read_sheet(excel_path, "Table 13", header_row=2)

    csr_factors: dict[int, float] = {}

    # Find columns
    indicator_col = None
    factor_col = None
    for col in df.columns:
        col_str = str(col).strip().lower()
        if "csr indicator" in col_str or "person-level" in col_str:
            indicator_col = col
        elif "csr" in col_str and "factor" in col_str:
            factor_col = col

    if indicator_col is None or factor_col is None:
        return _default_csr_factors()

    for _, row in df.iterrows():
        ind = row[indicator_col]
        factor = row[factor_col]
        if pd.notna(ind) and pd.notna(factor):
            ind = int(ind)
            csr_factors[ind] = float(factor)

    logger.info("Table 13: %d CSR indicator mappings", len(csr_factors))
    return csr_factors


def _default_csr_factors() -> dict[int, float]:
    """Default CSR indicator → factor mapping."""
    return {
        1: 1.00,
        2: 1.07,
        3: 1.12,
        4: 1.51,
        5: 1.19,
        6: 1.31,
        7: 1.04,
        8: 1.39,
        9: 1.10,
        10: 1.46,
        11: 1.15,
    }


def _parse_table12(excel_path: Path, sheet_name: str) -> dict[str, list[str]]:
    """Parse Table 12: Model Exclusions (informational only)."""
    df = _read_sheet(excel_path, sheet_name, header_row=2)

    exclusions: dict[str, list[str]] = {"Adult": [], "Child": [], "Infant": []}

    # Find columns for each model
    hcc_col = None
    adult_col = None
    child_col = None
    infant_col = None

    for col in df.columns:
        col_str = str(col).strip().lower()
        if "hcc" in col_str and "label" not in col_str and "excluded" not in col_str:
            hcc_col = col
        elif "adult" in col_str:
            adult_col = col
        elif "child" in col_str:
            child_col = col
        elif "infant" in col_str:
            infant_col = col

    if hcc_col is None:
        return exclusions

    for _, row in df.iterrows():
        hcc = row[hcc_col]
        if pd.isna(hcc):
            continue
        try:
            hcc_str = str(int(float(hcc))).zfill(3)
        except (ValueError, TypeError):
            continue
        if adult_col and pd.notna(row[adult_col]) and str(row[adult_col]).strip().upper() == "X":
            exclusions["Adult"].append(hcc_str)
        if child_col and pd.notna(row[child_col]) and str(row[child_col]).strip().upper() == "X":
            exclusions["Child"].append(hcc_str)
        if infant_col and pd.notna(row[infant_col]) and str(row[infant_col]).strip().upper() == "X":
            exclusions["Infant"].append(hcc_str)

    return exclusions


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_excel(excel_path: Path, benefit_year: int) -> ModelData:
    """Parse a CMS DIY Excel workbook into a ModelData object."""
    logger.info("Parsing %s for benefit year %d", excel_path, benefit_year)

    # Get sheet names
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    # Helper to find sheet name tolerating whitespace
    def find_sheet(target: str) -> str:
        for s in sheet_names:
            if s.strip() == target.strip():
                return s
        raise KeyError(f"Sheet '{target}' not found")

    # Parse each table
    icd10_cc = _parse_table3(excel_path, find_sheet("Table 3"), benefit_year)
    hcc_hierarchy, all_hccs = _parse_table4(excel_path, find_sheet("Table 4"))
    age_sex_vars = _parse_table5_age_sex(excel_path, find_sheet("Table 5"))

    # Parse variable tables for each model segment
    adult_vars = _parse_variable_table(excel_path, find_sheet("Table 6"), "Adult")
    child_vars = _parse_variable_table(excel_path, find_sheet("Table 7"), "Child")
    infant_vars = _parse_variable_table(excel_path, find_sheet("Table 8"), "Infant")

    # Merge HCC groups
    hcc_groups = {
        "Adult": adult_vars["hcc_groups"],
        "Child": child_vars["hcc_groups"],
    }

    # Merge severity/transplant triggers
    severe_triggers = {
        "Adult": adult_vars["severe_triggers"],
        "Child": child_vars["severe_triggers"],
    }
    transplant_triggers = {
        "Adult": adult_vars["transplant_triggers"],
        "Child": child_vars["transplant_triggers"],
    }

    # Merge severe/transplant HCC count bins
    severe_hcc_count_bins = {
        "Adult": adult_vars["severe_hcc_count_bins"],
        "Child": child_vars["severe_hcc_count_bins"],
    }
    transplant_hcc_count_bins = {
        "Adult": adult_vars["transplant_hcc_count_bins"],
        "Child": child_vars["transplant_hcc_count_bins"],
    }

    factors = _parse_table9(excel_path, find_sheet("Table 9"))
    ndc_rxc = _parse_table10a(excel_path, find_sheet("Table 10a"))
    hcpcs_rxc = _parse_table10b(excel_path, find_sheet("Table 10b"))
    rxc_hierarchy, all_rxcs = _parse_table11(excel_path, find_sheet("Table 11"))
    model_exclusions = _parse_table12(excel_path, find_sheet("Table 12"))
    csr_factors = _parse_table13(excel_path, sheet_names)

    # HCC_CNT exclusion
    hcc_cnt_exclude = {
        "Adult": adult_vars["hcc_cnt_exclude"],
        "Child": child_vars.get("hcc_cnt_exclude"),
    }

    return ModelData(
        benefit_year=benefit_year,
        icd10_cc=icd10_cc,
        hcc_hierarchy=hcc_hierarchy,
        all_hccs=all_hccs,
        age_sex_vars=age_sex_vars,
        hcc_groups=hcc_groups,
        rxc_interactions=adult_vars["rxc_interactions"],
        severe_triggers=severe_triggers,
        transplant_triggers=transplant_triggers,
        severe_hcc_count_bins=severe_hcc_count_bins,
        transplant_hcc_count_bins=transplant_hcc_count_bins,
        enrollment_duration_vars=adult_vars["enrollment_duration_vars"],
        infant_severity_levels=infant_vars["infant_severity_levels"],
        infant_maturity_categories=infant_vars["infant_maturity_categories"],
        infant_maturity_severity_interactions=infant_vars["infant_maturity_severity_interactions"],
        factors=factors,
        ndc_rxc=ndc_rxc,
        hcpcs_rxc=hcpcs_rxc,
        rxc_hierarchy=rxc_hierarchy,
        all_rxcs=all_rxcs,
        model_exclusions=model_exclusions,
        csr_factors=csr_factors,
        hcc_cnt_exclude=hcc_cnt_exclude,
    )
